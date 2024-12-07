import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from src.utils.logger_setup import logger

def save_data_to_excel(data, league_name):
    """
    Сохраняет данные в Excel-файл для указанной лиги. Создает новый файл или обновляет существующий.
    :param data: Словарь данных матча или список словарей.
    :param league_name: Название лиги (используется для имени файла).
    """
    try:
        output_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "data"))
        os.makedirs(output_dir, exist_ok=True)
        
        file_path = os.path.join(output_dir, f"{league_name}.xlsx")

        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
            existing_rows = sheet.max_row
            logger.info(f"Файл {file_path} найден. Данные будут добавлены.")
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = league_name
            existing_rows = 0
            _add_headers(sheet)
            logger.info("Создан новый файл с заголовками столбцов.")

        if isinstance(data, dict):
            data = [data]

        for lot in data:
            _write_row(sheet, existing_rows + 1, lot)
            existing_rows += 1

        _adjust_column_widths(sheet, existing_rows)

        workbook.save(file_path)
        logger.info(f"Данные успешно сохранены в файл: {file_path}")

    except Exception as e:
        logger.error(f"Ошибка при сохранении данных в файл {league_name}.xlsx: {e}")

def _add_headers(sheet):
    """
    Добавляет заголовки в лист Excel.
    :param sheet: Лист Excel.
    """
    headers = [
        "Team (Home)", "Team (Away)", "Home / Away", "Over / Under", "Both To Score",
        "Correct Score", "Team Rating (Home)", "Team Rating (Away)", "Team Form (Home)",
        "Team Form (Away)", "XG Luckiness (Home)", "XG Luckiness (Away)", "XG Predictability (Home)",
        "XG Predictability (Away)", "Avg XG Scored (Home)", "Avg XG Scored (Away)",
        "Avg XG Conceded (Home)", "Avg XG Conceded (Away)", "Match Score Prediction (Home)",
        "Match Score Prediction (Away)", "Goals (Home)", "Goals (Away)",
        "Expected Goals (Home)", "Expected Goals (Away)"
    ]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _write_row(sheet, row_num, match):
    """
    Записывает одну строку данных в лист Excel.
    :param sheet: Лист Excel.
    :param row_num: Номер строки.
    :param match: Словарь с данными матча.
    """
    row = [
        match["preview"].get("team_name_1", ""),
        match["preview"].get("team_name_2", ""),
        match["preview"].get("winner", ""),
        match["preview"].get("total_under", ""),
        match["preview"].get("both_to_score", ""),
        match["preview"].get("correct_score", ""),
        match["preview"].get("team_rating_home", ""),
        match["preview"].get("team_rating_away", ""),
        match["preview"].get("team_form_home", ""),
        match["preview"].get("team_form_away", ""),
        match["preview"].get("xg_luckiness_home", ""),
        match["preview"].get("xg_luckiness_away", ""),
        match["preview"].get("xg_predictability_home", ""),
        match["preview"].get("xg_predictability_away", ""),
        match["preview"].get("avg_xg_scored_home", ""),
        match["preview"].get("avg_xg_scored_away", ""),
        match["preview"].get("avg_xg_conceded_home", ""),
        match["preview"].get("avg_xg_conceded_away", ""),
        match["match_score_prediction"].get("match_score_prediction_home", ""),
        match["match_score_prediction"].get("match_score_prediction_away", ""),
        match["xg_statistics"].get("goals_team_1", ""),
        match["xg_statistics"].get("goals_team_2", ""),
        match["xg_statistics"].get("expected_goals_team_1", ""),
        match["xg_statistics"].get("expected_goals_team_2", ""),
    ]
    for col_num, value in enumerate(row, start=1):
        cell = sheet.cell(row=row_num + 1, column=col_num, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _adjust_column_widths(sheet, up_to_row):
    """
    Корректирует ширину столбцов на основе данных.
    :param sheet: Лист Excel.
    :param up_to_row: Количество строк, до которых корректируется ширина.
    """
    for col_num in range(1, sheet.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for row in range(1, up_to_row + 1):
            cell = sheet.cell(row=row, column=col_num)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width
